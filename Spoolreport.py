import datetime
import io
import os
import requests
from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.colors import HexColor
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
import streamlit as st

# -------- CONFIGURATION --------
GOOGLE_SHEET_URL = "https://docs.google.com/spreadsheets/d/1P1-U_1rhYJ28drrdGgwKBVntP9Uh4nlQ/edit?usp=sharing"

st.set_page_config(page_title="Spool Detail Report Generator", layout="wide")
st.title("📊 Spool Detail Report Generator")


# -------- GOOGLE SHEET FETCHER (Openpyxl Live) --------
def get_web_workbook(url):
    try:
        if "spreadsheets/d/" in url:
            file_id = url.split("spreadsheets/d/")[1].split("/")[0]
        elif "id=" in url:
            file_id = url.split("id=")[1].split("&")[0]
        else:
            st.error("❌ લિંક ફોર્મેટ ખોટું છે.")
            return None

        d_url = f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=xlsx"
        response = requests.get(d_url)
        if response.status_code == 200:
            return load_workbook(
                filename=io.BytesIO(response.content), data_only=True
            )
        else:
            st.error("❌ ફાઇલ એક્સેસ નકારી.")
            return None
    except Exception as e:
        st.error(f"❌ એરર: {e}")
        return None


# -------- HELPERS --------
def clean_val(x):
    if x is None or str(x).strip().lower() in ["na", "nan", "none", ""]:
        return ""
    if isinstance(x, (datetime.datetime, datetime.date)):
        return x.strftime("%d-%m-%Y")
    return str(x).strip()


# -------- MAIN WEB APP LOGIC --------
with st.spinner("⏳ ગૂગલ શીટમાંથી લાઈવ ડેટા લોડ થઈ રહ્યો છે..."):
    wb = get_web_workbook(GOOGLE_SHEET_URL)

if wb is not None:
    if "Sheet2" not in wb.sheetnames:
        st.error("❌ Sheet2 મળી નથી.")
    else:
        ws = wb["Sheet2"]

        # સાઇડબાર સર્ચ
        st.sidebar.header("🔍 સર્ચ પેનલ")
        usr_no = st.sidebar.text_input(
            "Spool Unique No. લખો:", placeholder="e.g., A-41101"
        ).strip()

        if usr_no:
            filtered_rows = []
            # સેમ લોજિક: રો ૨ થી છેલ્લી રો સુધી લૂપ ફેરવવી
            for r in range(2, ws.max_row + 1):
                # કોલમ G માં Spool Unique No છે
                if clean_val(ws[f"G{r}"].value) == usr_no:
                    filtered_rows.append(r)

            if not filtered_rows:
                st.warning(f"⚠️ {usr_no} માટે કોઈ રેકોર્ડ મળ્યો નથી.")
            else:
                st.success(f"✅ {len(filtered_rows)} રેકોર્ડ્સ મળ્યા!")

                # સ્ક્રીન પ્રીવ્યૂ માટે લિસ્ટ તૈયાર કરવું
                preview_data = []
                for r in filtered_rows:
                    preview_data.append(
                        {
                            "ISO/Drawing No": clean_val(ws[f"F{r}"].value),
                            "Joint No.": clean_val(ws[f"R{r}"].value),
                            "Type of Joint": clean_val(ws[f"J{r}"].value),
                            "WELD NPD": clean_val(ws[f"H{r}"].value),
                            "Spool Unique No.": clean_val(ws[f"G{r}"].value),
                            "FIT UP Date": clean_val(ws[f"AA{r}"].value),
                            "Welder No": clean_val(ws[f"X{r}"].value),
                            "VISUAL Date": clean_val(ws[f"AB{r}"].value),
                        }
                    )

                st.subheader("📋 લાઈવ ડેટા પ્રીવ્યૂ")
                st.dataframe(pd.DataFrame(preview_data), use_container_width=True)

                # -------- PDF BUILDER (Using Openpyxl Data) --------
                def generate_pdf_bytes():
                    buffer = io.BytesIO()
                    doc = SimpleDocTemplate(
                        buffer,
                        pagesize=landscape(A4),
                        rightMargin=15,
                        leftMargin=15,
                        topMargin=15,
                        bottomMargin=15,
                    )

                    elements = []
                    title_style = ParagraphStyle(
                        "Title",
                        fontSize=14,
                        textColor=HexColor("#2C3E50"),
                        alignment=1,
                        spaceAfter=12,
                    )
                    body_style = ParagraphStyle(
                        "Body", fontSize=6.5, leading=9, alignment=1
                    )
                    header_style = ParagraphStyle(
                        "Header",
                        fontSize=7,
                        leading=8,
                        alignment=1,
                        textColor=colors.whitesmoke,
                        fontName="Helvetica-Bold",
                    )

                    elements.append(
                        Paragraph(
                            f"SPOOL DETAIL REPORT - {usr_no}", title_style
                        )
                    )

                    headers = [
                        "ISO No/Drawing No/Line No",
                        "Joint No.",
                        "Type of Joint",
                        "WELD NPD",
                        "Spool Unique No.",
                        "FIT UP Date",
                        "Welder No",
                        "VISUAL Date",
                    ]
                    table_data = [[Paragraph(h, header_style) for h in headers]]

                    for r in filtered_rows:
                        row_cells = [
                            clean_val(ws[f"F{r}"].value),
                            clean_val(ws[f"R{r}"].value),
                            clean_val(ws[f"J{r}"].value),
                            clean_val(ws[f"H{r}"].value),
                            clean_val(ws[f"G{r}"].value),
                            clean_val(ws[f"AA{r}"].value),
                            clean_val(ws[f"X{r}"].value),
                            clean_val(ws[f"AB{r}"].value),
                        ]
                        table_data.append(
                            [
                                Paragraph(val if val else "-", body_style)
                                for val in row_cells
                            ]
                        )

                    column_widths = [130, 42, 42, 42, 65, 60, 42, 60]
                    table = Table(
                        table_data,
                        colWidths=column_widths,
                        rowHeights=[40] * len(table_data),
                        repeatRows=1,
                    )
                    table.setStyle(
                        TableStyle(
                            [
                                (
                                    "BACKGROUND",
                                    (0, 0),
                                    (-1, 0),
                                    HexColor("#2C3E50"),
                                ),
                                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                            ]
                        )
                    )
                    elements.append(table)

                    # એક્સ્ટ્રા માહિતિ (પ્રથમ મેચિંગ રો માંથી)
                    f_row = filtered_rows[0]
                    ndt_clearance = clean_val(ws[f"AR{f_row}"].value)

                    elements.append(Spacer(1, 15))
                    extra_style = ParagraphStyle(
                        "ExtraStyle",
                        fontSize=9,
                        fontName="Helvetica-Bold",
                        leading=12,
                    )
                    elements.append(
                        Paragraph(
                            f"NDT CLEARANCE / FD DATE:- {ndt_clearance}",
                            extra_style,
                        )
                    )

                    doc.build(elements)
                    buffer.seek(0)
                    return buffer

                # PDF ડાઉનલોડ બટન
                st.sidebar.markdown("---")
                st.sidebar.subheader("📥 રીપોર્ટ ડાઉનલોડ")

                pdf_data = generate_pdf_bytes()
                st.sidebar.download_button(
                    label="📥 Download PDF Report",
                    data=pdf_data,
                    file_name=f"Spool_Report_{usr_no}.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                )
